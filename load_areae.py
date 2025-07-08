
import pandas as pd
import openpyxl
from sqlalchemy import text
import urllib
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError

import pyodbc
import pandas as pd
from datetime import datetime, timedelta, date
from sqlalchemy import create_engine
import urllib
import argparse
import schedule
import time
import pandas as pd
import os
import pandas as pd
import numpy as np
# import schedule
import pandas as pd
# import xlsxwriter
import time
import pandas as pd
import pymongo
from pymongo import MongoClient
from pymongo import MongoClient
from datetime import datetime, timedelta
from sqlalchemy.orm import sessionmaker
import logging
from sqlalchemy.exc import IntegrityError
import logging
import os
from typing import List, Dict

from sqlalchemy.engine import Engine





def read_config(filename='config.txt'):
    """
    Reads a key-value configuration from a text file.

    :param filename: Name of the configuration file. Default is 'config.txt'.
    :return: A dictionary containing configuration key-value pairs.
    """
    script_path = os.path.abspath(__file__)
    config_path = os.path.join(os.path.dirname(script_path), filename)
    config = {}

    # Verify that the file exists
    if not os.path.exists(config_path):
        print(f"Error: Configuration file '{filename}' not found at {config_path}.")
        return {}

    try:
        with open(config_path, 'r') as file:
            for line in file:
                line = line.strip()
                # Skip empty lines or comments
                if not line or line.startswith('#'):
                    continue
                try:
                    key, value = line.split('=', 1)  # Limit split to at most 2 items
                    config[key.strip()] = value.strip().strip("'\"")  # Remove quotes
                except ValueError:
                    print(f"Warning: Skipping malformed line: {line}")
    except Exception as e:
        print(f"Error reading configuration file: {e}")
        return {}

    return config


# Get configuration values
config = read_config()

# Access individual parameters
excel_prod = config.get('excel_prod', '')
excel_test = config.get('excel_test', '')
environment = config.get('environment', '')
days_ago = 7
# Connection string
# Connection string
conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=lulimdbserver.database.windows.net;'
    'DATABASE=lulimdb;'
    'UID=shmuelstav;'
    'PWD=5tgbgt5@;'
    'TrustServerCertificate=yes;'
    'Encrypt=yes;'
)

# URL encode the connection string
params = urllib.parse.quote_plus(conn_str)

# Create SQLAlchemy engine
engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}', echo=False)


# Print or use the values as needed
print("Excel Prod:", excel_prod)
print("Excel Test:", excel_test)
print("Environment:", environment)

farms = '\\fandy farms'
reports = '\\current active flocks\\'
sheet_name_tmuta = 'tmuta'
sheet_name_sivuk = 'shivuk'
sheet_name_skila = 'סיכום שקילות'
sheet_name_mivne = 'שטח מבנים'
sheet_name_tarovet = 'תערובת'
excel_file_name_finish = 'current flock '
excel_end = '.xlsx'''
excel_middle_name = '\\current flock\\'

farms_new_folk = {}


# Functions
def read_excel(path, sheet_name):
    try:
        # df = pd.read_excel(path, sheet_name=sheet_name)
        # return df
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # Check if the specified sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found.")
            return None

        # Access the specified sheet even if it's hidden
        sheet = workbook[sheet_name]

        # Read the sheet data into a DataFrame
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=data[0])

        # Close the workbook
        workbook.close()

        return df

    except Exception as e:
        # Handle any exceptions (e.g., file not found, sheet not found) and return an empty DataFrame
        print(f"An error occurred: {str(e)}")
        return pd.DataFrame()


def subfolder_names(path):
    folder_names = []

    # List all directories and files in the given path
    entries = os.listdir(path)

    for entry in entries:
        entry_path = os.path.join(path, entry)

        # Check if the entry is a directory (subfolder)
        if os.path.isdir(entry_path):
            folder_names.append(entry)

    return folder_names


def translate(farm):
    translations = {
        'gotliv 4': 'גוטליב 4',
        'gotliv 2': 'גוטליב 2',
        'megadim': 'מגדים',
        'megido': 'מגידו',
        'ranen': 'רנן',
        'shaal morad': 'שעל - מורד',
        'kalanit': 'כלנית',
        'ramat zvi haim': 'רמת צבי חיים',
        'ramat zvi moshe': 'רמת צבי משה',
        'ramot naftali': 'רמות נפתלי',
        'ranen': 'רנן',
        'shaal moyal': 'שעל - מויאל',
        'musayel': 'מוסייל',
        'sigler': 'סיגלר',
        'gazit': 'גזית',
        'sadmot dvora': 'שדמות דבורה',
        'mawiya': 'מעאוויה',
        'sharona': 'שרונה'
        # Add more translations as needed
    }

    # Check if farm exists in translations dictionary
    if farm in translations:
        return translations[farm]
    else:
        print(f"Translation not found for '{farm}'")
        return f"Translation not found for '{farm}'"



def clean_and_filter_data(df):
    if df is None or df.empty:
        return pd.DataFrame()

    # שלב 1: הסר שורות ועמודות ריקות לגמרי
    df = df.dropna(how='all').dropna(axis=1, how='all')

    # שלב 2: אפס אינדקס
    df = df.reset_index(drop=True)

    # שלב 3: קח רק את שתי העמודות הראשונות
    df = df.iloc[:, :2]

    # שלב 3.5: הפוך את השורה הראשונה לכותרות
    df.columns = df.iloc[0]
    df = df[1:]  # מחק את השורה הראשונה כי היא הפכה לכותרת

    # שלב 4: סנן רק שורות שבהן הערך בעמודה הראשונה מתחיל במספר
    first_col = df.columns[0]
    df = df[df[first_col].astype(str).str.match(r'^\d')]

    # אפס את האינדקס שוב לאחר הסינון
    df = df.reset_index(drop=True)

    return df

def prepare_rows_for_db(df, farm_name):
    if df is None or df.empty:
        return []

    # נניח שיש לך פונקציית תרגום
    translated_farm = translate(farm_name)

    # הוסף עמודת שם חווה
    df['farm_name'] = translated_farm

    # המרה לרשימת מילונים (dict) – מבנה שמתאים ל-SQLAlchemy או הכנסת DB אחרת
    rows = df.to_dict(orient='records')

    return rows


def transform_and_prepare_for_db(rows_list):
    import pandas as pd

    if not rows_list:
        return []

    # המרה ל-DataFrame
    df = pd.DataFrame(rows_list)

    # מפת שמות עמודות
    rename_map = {
        "שם חווה": "farm_name",
        "מס' מבנה": "mivne",
        'מ"ר': "area"
    }

    # החלפת שמות עמודות (רק אם קיימים)
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    # החזרת רשימה של dicts
    return df.to_dict(orient='records')


def insert_if_not_exists(engine: Engine, rows_list: List[Dict[str, any]]) -> None:
    """
    מבצע INSERT לתוך dbo.farms_structures רק אם הרשומה לא קיימת כבר.

    :param engine: אובייקט SQLAlchemy Engine
    :param rows_list: רשימה של מילונים בפורמט:
                      [{"farm_name": ..., "mivne": ..., "area": ...}, ...]
    """
    if not rows_list:
        print("🔹 No data to insert. rows_list is empty.")
        return

    # בדיקה שכל השדות הדרושים קיימים בכל שורה
    required_keys = {"farm_name", "mivne", "area"}
    for i, row in enumerate(rows_list):
        if not required_keys.issubset(row.keys()):
            raise ValueError(f"❌ Row {i} is missing required keys: {required_keys - row.keys()}")

    insert_sql = text("""
        INSERT INTO dbo.farms_structures (farm_name, mivne, area)
        SELECT :farm_name, :mivne, :area
        WHERE NOT EXISTS (
            SELECT 1 FROM dbo.farms_structures
            WHERE farm_name = :farm_name AND mivne = :mivne AND area = :area
        )
    """)

    try:
        with engine.begin() as conn:  # מנהל טרנזקציה
            conn.execute(insert_sql, rows_list)
            print(f"✅ Inserted {len(rows_list)} rows if not already existing.")
    except Exception as e:
        print(f"❌ Failed to insert rows: {e}")
        raise

def run_load_area():
    farms_names = subfolder_names(excel_prod + farms)
    tmuta_results = pd.DataFrame()
    for farm in farms_names:

        path = f"{excel_prod}{farms}\\{farm}{excel_middle_name}{excel_file_name_finish}{farm}{excel_end}"
        # path = 'C:\\Users\\User\\Dropbox\\BMC\\prod\\fandy farms\\shaal moyal\\current flock\\current flock shaal moyal.xlsx'
        data = read_excel(path, sheet_name_mivne)

        if not data.empty:
            df_end =  clean_and_filter_data(data)
            df_to_df = prepare_rows_for_db(df_end,farm)
            df_to_df = transform_and_prepare_for_db(df_to_df)
            insert_if_not_exists(engine, df_to_df)




def run_file():
    # הגדרת מחרוזת החיבור
    conn_str = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=lulimdbserver.database.windows.net;'
        'DATABASE=lulimdb;'
        'UID=shmuelstav;'
        'PWD=5tgbgt5@;'
        'TrustServerCertificate=yes;'
        'Encrypt=yes;'
    )
    params = urllib.parse.quote_plus(conn_str)
    engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}', fast_executemany=True)

    # קריאת קובץ האקסל
    df = pd.read_excel("C:/Users/User/Documents/farm_area.xlsx")
    df = df[['farm_name', 'mivne', 'area']].dropna()

    # ניקוי שדה farm_name — שומר אותיות עבריות, אנגליות, ספרות, רווחים ומקפים
    df['farm_name'] = df['farm_name'].astype(str).str.strip().str.replace(r'[^א-תa-zA-Z0-9\s-]', '', regex=True)

    # המרת טיפוסים ל-int
    df['mivne'] = df['mivne'].astype(int)
    df['area'] = df['area'].astype(int)

    with engine.begin() as conn:
        # מחיקת טבלה זמנית במידה וקיימת ויצירתה מחדש
        conn.execute(text("""
            IF OBJECT_ID('tempdb..#temp_farms_structures') IS NOT NULL DROP TABLE #temp_farms_structures;

            CREATE TABLE #temp_farms_structures (
                farm_name NVARCHAR(50) COLLATE Hebrew_CI_AS NOT NULL,
                mivne INT NOT NULL,
                area INT NOT NULL
            );
        """))

        # הוספת שורות לטבלה הזמנית אחת-אחת
        insert_sql = text("""
            INSERT INTO #temp_farms_structures (farm_name, mivne, area)
            VALUES (:farm_name, :mivne, :area)
        """)
        for _, row in df.iterrows():
            conn.execute(insert_sql, {
                "farm_name": row['farm_name'],
                "mivne": row['mivne'],
                "area": row['area']
            })

        # הרצת MERGE להוספת רשומות חדשות לטבלה הראשית בלבד
        merge_sql = """
            MERGE dbo.farms_structures AS target
            USING #temp_farms_structures AS source
            ON 
                target.farm_name COLLATE Hebrew_CI_AS = source.farm_name COLLATE Hebrew_CI_AS
                AND target.mivne = source.mivne
                AND target.area = source.area
            WHEN NOT MATCHED BY TARGET THEN
                INSERT (farm_name, mivne, area)
                VALUES (source.farm_name, source.mivne, source.area);
        """
        result = conn.execute(text(merge_sql))
        print("✔️ Merge completed without duplicates.")
        print(f"{result.rowcount} rows inserted.")


if __name__ == "__main__":
    # Execute the main function when the script is run independently.
    run_file()
    #run_load_area()
