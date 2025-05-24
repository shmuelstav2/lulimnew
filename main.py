# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import pandas as pd
import openpyxl
from sqlalchemy import text
import urllib
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError

import pyodbc
import pandas as pd
from datetime import datetime, timedelta, date
from sqlalchemy import create_engine
import urllib
import argparse
import schedule
import time
import pandas as pd
import os
import pandas as pd
import numpy as np
# import schedule
import pandas as pd
# import xlsxwriter
import time
import pandas as pd
import pymongo
from pymongo import MongoClient
from pymongo import MongoClient
from datetime import datetime, timedelta
from sqlalchemy.orm import sessionmaker
import logging
from sqlalchemy.exc import IntegrityError


import logging
import os

# וודא תיקיית לוגים
os.makedirs('logs', exist_ok=True)

logging.basicConfig(
    filename='logs/sivuk_errors.log',
    filemode='a',
    format='%(asctime)s %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.ERROR
)
# Configure logging
logging.basicConfig(level=logging.INFO)  # Set the default logging level to INFO

# Suppress SQLAlchemy logs
logging.getLogger('sqlalchemy.engine').setLevel(logging.WARNING)  # Only show warnings or errors for SQLAlchemy

import warnings
warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)


import warnings

# Suppress warnings from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Suppress pandas warnings
warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

class NoSQLAlchemyLogs(logging.Filter):
    def filter(self, record):
        return 'sqlalchemy' not in record.name

# Apply the filter to the root logger
for handler in logging.root.handlers:
    handler.addFilter(NoSQLAlchemyLogs())


def read_config(filename='config.txt'):
    """
    Reads a key-value configuration from a text file.

    :param filename: Name of the configuration file. Default is 'config.txt'.
    :return: A dictionary containing configuration key-value pairs.
    """
    script_path = os.path.abspath(__file__)
    config_path = os.path.join(os.path.dirname(script_path), filename)
    config = {}

    # Verify that the file exists
    if not os.path.exists(config_path):
        print(f"Error: Configuration file '{filename}' not found at {config_path}.")
        return {}

    try:
        with open(config_path, 'r') as file:
            for line in file:
                line = line.strip()
                # Skip empty lines or comments
                if not line or line.startswith('#'):
                    continue
                try:
                    key, value = line.split('=', 1)  # Limit split to at most 2 items
                    config[key.strip()] = value.strip().strip("'\"")  # Remove quotes
                except ValueError:
                    print(f"Warning: Skipping malformed line: {line}")
    except Exception as e:
        print(f"Error reading configuration file: {e}")
        return {}

    return config


# Get configuration values
config = read_config()

# Access individual parameters
excel_prod = config.get('excel_prod', '')
excel_test = config.get('excel_test', '')
environment = config.get('environment', '')
days_ago = 7
# Connection string
# Connection string
conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=lulimdbserver.database.windows.net;'
    'DATABASE=lulimdb;'
    'UID=shmuelstav;'
    'PWD=5tgbgt5@;'
    'TrustServerCertificate=yes;'
    'Encrypt=yes;'
)

# URL encode the connection string
params = urllib.parse.quote_plus(conn_str)

# Create SQLAlchemy engine
engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}', echo=False)


# Print or use the values as needed
print("Excel Prod:", excel_prod)
print("Excel Test:", excel_test)
print("Environment:", environment)

farms = '\\fandy farms'
reports = '\\current active flocks\\'
sheet_name_tmuta = 'tmuta'
sheet_name_sivuk = 'shivuk'
sheet_name_skila = 'סיכום שקילות'
sheet_name_tarovet = 'תערובת'
excel_file_name_finish = 'current flock '
excel_end = '.xlsx'''
excel_middle_name = '\\current flock\\'

farms_new_folk = {}


# Functions
def read_excel(path, sheet_name):
    try:
        # df = pd.read_excel(path, sheet_name=sheet_name)
        # return df
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # Check if the specified sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found.")
            return None

        # Access the specified sheet even if it's hidden
        sheet = workbook[sheet_name]

        # Read the sheet data into a DataFrame
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=data[0])

        # Close the workbook
        workbook.close()

        return df

    except Exception as e:
        # Handle any exceptions (e.g., file not found, sheet not found) and return an empty DataFrame
        print(f"An error occurred: {str(e)}")
        return pd.DataFrame()


def subfolder_names(path):
    folder_names = []

    # List all directories and files in the given path
    entries = os.listdir(path)

    for entry in entries:
        entry_path = os.path.join(path, entry)

        # Check if the entry is a directory (subfolder)
        if os.path.isdir(entry_path):
            folder_names.append(entry)

    return folder_names


def translate(farm):
    translations = {
        'gotliv 4': 'גוטליב 4',
        'gotliv 2': 'גוטליב 2',
        'megadim': 'מגדים',
        'megido': 'מגידו',
        'ranen': 'רנן',
        'shaal morad': 'שעל - מורד',
        'kalanit': 'כלנית',
        'ramat zvi haim': 'רמת צבי חיים',
        'ramat zvi moshe': 'רמת צבי משה',
        'ramot naftali': 'רמות נפתלי',
        'ranen': 'רנן',
        'shaal moyal': 'שעל - מויאל',
        'musayel': 'מוסייל',
        'sigler': 'סיגלר',
        'gazit': 'גזית',
        'sadmot dvora': 'שדמות דבורה',
        'mawiya': 'מעאוויה',
        'sharona': 'שרונה'
        # Add more translations as needed
    }

    # Check if farm exists in translations dictionary
    if farm in translations:
        return translations[farm]
    else:
        print(f"Translation not found for '{farm}'")
        return f"Translation not found for '{farm}'"


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def find_hathala(farm, day, df):
    df = df[df['farm name'] == farm]
    df = df.groupby('house number').agg({'mixed start quantity': 'max'}).reset_index()
    column_sum = df['mixed start quantity'].sum()
    return column_sum


def find_tmuta_iomit(farm, day, df):
    df = df[df['farm name'] == farm]
    df = df[df['growth day'] == day]
    column_sum = df['mixed daily mortality'].sum()
    return column_sum


def write_to_mongo_and_delete(df, db_name, collection_name):
    # MongoDB connection settings
    mongo_uri = "mongodb://localhost:27017/"
    database_name = db_name
    collection_name = collection_name

    # Convert DataFrame to a list of dictionaries
    data_list = df.to_dict(orient='records')

    # Convert datetime.date objects to datetime.datetime objects

    for record in data_list:
        for key, value in record.items():
            if isinstance(value, date):
                record[key] = value.strftime('%Y-%m-%d')
    # Connect to MongoDB
    client = MongoClient(mongo_uri)

    # Access the lulim database (it will be created if it doesn't exist)
    db = client[database_name]

    # Access the tmuta_end collection
    collection = db[collection_name]

    # Delete existing data in the collection
    collection.delete_many({})

    # Insert new data into the collection
    collection.insert_many(data_list)

    # Close the MongoDB connection
    client.close()


def udate_skila():
    farms_names = subfolder_names(excel_prod + farms)
    tmuta_results = pd.DataFrame()
    for farm in farms_names:
        path = f"{excel_prod}{farms}\\{farm}{excel_middle_name}{excel_file_name_finish}{farm}{excel_end}"
        # path = 'C:\\Users\\User\\Dropbox\\BMC\\prod\\fandy farms\\shaal moyal\\current flock\\current flock shaal moyal.xlsx'
        data = read_excel(path, sheet_name_skila)

        if not data.empty:
            data = data.replace('', np.nan)

            # Drop rows and columns where all elements are NaN
            df_cleaned = data.dropna(how='all').dropna(axis=1, how='all')

            columns_to_keep = [0, 1, 6, 7]
            df_cleaned = df_cleaned.iloc[:, columns_to_keep].dropna()

            new_columns = df_cleaned.iloc[0]
            df_cleaned = df_cleaned[1:]
            df_cleaned.columns = new_columns
            df_cleaned = df_cleaned.reset_index(drop=True)

            # Create new DataFrame for SQL
            df = pd.DataFrame()
            df['grotwh_day'] = pd.to_numeric(df_cleaned['יום'])
            df['avg_mixed'] = pd.to_numeric(df_cleaned['ממוצע מעורב'])
            df['avg_mixed_percent'] = pd.to_numeric(df_cleaned['אחוז תקן מעורב'])
            df['mivne'] = pd.to_numeric(df_cleaned['מבנה'])
            if not df.empty:
                new_flock = 'new_flock'  # Define the new flock column name
                df[new_flock] = farms_new_folk[farm]
                # df['midgar'] = 1
                df['farm_name'] = str(translate(farm))
                df['avg_mixed'] = df['avg_mixed'].round(3)
                df['avg_mixed_percent'] = df['avg_mixed_percent'].round(3)
                df = df[~((df['avg_mixed'] == 0) & (df['avg_mixed_percent'] == 0))]

                # Load existing data and normalize
                existing_data = pd.read_sql('SELECT grotwh_day, mivne, new_flock, farm_name FROM skila_svuit',
                                            con=engine)
                existing_data = existing_data.astype(
                    {'grotwh_day': 'int64', 'mivne': 'int64', 'new_flock': 'int64', 'farm_name': 'str'})
                existing_data.set_index(['grotwh_day', 'mivne', 'new_flock', 'farm_name'], inplace=True)

                df = df.astype({'grotwh_day': 'int64', 'mivne': 'int64', 'new_flock': 'int64', 'farm_name': 'str'})
                df.set_index(['grotwh_day', 'mivne', 'new_flock', 'farm_name'], inplace=True)

                # Identify new rows
                new_rows = df[~df.index.isin(existing_data.index)].reset_index()

                # Upsert logic: Insert new rows and update existing ones
                if not new_rows.empty:
                    with engine.begin() as connection:
                        for _, row in new_rows.iterrows():
                            stmt = text("""
                            MERGE INTO skila_svuit AS target
                            USING (SELECT :grotwh_day AS grotwh_day, :mivne AS mivne, :new_flock AS new_flock, :farm_name AS farm_name) AS source
                            ON target.grotwh_day = source.grotwh_day AND target.mivne = source.mivne AND target.new_flock = source.new_flock AND target.farm_name = source.farm_name
                            WHEN MATCHED THEN
                                UPDATE SET avg_mixed = :avg_mixed, avg_mixed_percent = :avg_mixed_percent
                            WHEN NOT MATCHED THEN
                                INSERT (grotwh_day, mivne, new_flock, farm_name, avg_mixed, avg_mixed_percent)
                                VALUES (:grotwh_day, :mivne, :new_flock, :farm_name, :avg_mixed, :avg_mixed_percent);
                            """)
                            params = {
                                'grotwh_day': row['grotwh_day'],
                                'mivne': row['mivne'],
                                'new_flock': row['new_flock'],
                                'farm_name': row['farm_name'],
                                'avg_mixed': row['avg_mixed'],
                                'avg_mixed_percent': row['avg_mixed_percent']
                            }
                            connection.execute(stmt, params)  # Use params dictionary
                    print(f"Upserted rows successfully for {farm}.")

                else:
                    print(f"No new rows to upsert for {farm}.")

    # Fetch data and write to MongoDB
    df_view = pd.read_sql("""
        SELECT [farm_name], [mivne], [grotwh_day], [avg_mixed], [avg_mixed_percent], [parent_flock]
        FROM dbo.skila_svuit_highest_grotwh_day OPTION (RECOMPILE);
    """, con=engine)

    #df_view = pd.read_sql("SELECT * FROM dbo.skila_svuit_highest_grotwh_day OPTION (RECOMPILE);", con=engine)
    write_to_mongo_and_delete(df_view, 'lulim_new', 'skila')
    print('בכרבר')


def truncate_flock():
    conn_str = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=lulimdbserver.database.windows.net;'
        'DATABASE=lulimdb;'
        'UID=shmuelstav;'
        'PWD=5tgbgt5@'
    )

    # URL encoding for connection string
    params = urllib.parse.quote_plus(conn_str)
    engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}')
    try:
        # Use `engine.connect()` for a connection and manage transactions explicitly
        with engine.connect() as connection:
        #with engine.connect() as connection:
            # Use `connection.begin()` for transactional safety if required
            with connection.begin():
                # Execute the TRUNCATE TABLE statement
                stmt = text("TRUNCATE TABLE [dbo].[open_farms]")
                connection.execute(stmt)
                print("Table [dbo].[open_farms] successfully truncated.")
    except Exception as e:
        print(f"Error while truncating the table: {e}")

def add_flock(farm):
    with engine.begin() as connection:
        stmt = text("""
                           INSERT INTO [dbo].[open_farms] (farm_name) 
                          VALUES (:farm_name)
                       """)
        connection.execute(stmt, {'farm_name': farm})


def udate_tmuta():
    # run over all the farm files

    truncate_flock()

    farms_names = subfolder_names(excel_prod + farms)
    count = 1
    for farm in farms_names:

        path = f"{excel_prod}{farms}\\{farm}{excel_middle_name}{excel_file_name_finish}{farm}{excel_end}"
        data = read_excel(path, sheet_name_tmuta)
        if not data.empty:
            # and  count == 1:
            count = 2
            data = data.replace('', np.nan)

            # Drop rows and columns where all elements are NaN
            df_cleaned = data.dropna(how='all').dropna(axis=1, how='all')
            if not df_cleaned.empty:
                threshold = 5
                # Delete columns with fewer non-null values than the threshold
                df_cleaned = df_cleaned.dropna(axis=1, thresh=threshold)
                df_cleaned = df_cleaned.dropna(axis=0, thresh=threshold)
                df_cleaned.columns = df_cleaned.iloc[0]
                df_cleaned = df_cleaned.iloc[1:]
                df_cleaned = df_cleaned.reset_index(drop=True)
                df_cleaned = df_cleaned[df_cleaned['mixed daily mortality'] > 0]

            if not df_cleaned.empty:
                # Create new DataFrame for SQL
                df = pd.DataFrame()
                df_cleaned['site'] = df_cleaned['site'].astype(str)
                df_cleaned['site'] = df_cleaned['site'].str.split('.').str[0]
                df_cleaned['site'] = pd.to_numeric(df_cleaned['site'])
                df_cleaned['site'] = pd.to_numeric(df_cleaned['site'])
                df_cleaned['house_number'] = pd.to_numeric(df_cleaned['house number'])
                df_cleaned['perent flock'] = df_cleaned['perent flock'].astype(str)
                df_cleaned['parent_flock'] = df_cleaned['perent flock'].str.extract('(\d+)')
                df_cleaned['parent_flock'] = pd.to_numeric(df_cleaned['parent_flock'])
                df_cleaned['mixed_start'] = pd.to_numeric(df_cleaned['mixed start quantity'])
                df_cleaned['daily_mortality'] = pd.to_numeric(df_cleaned['daily mortality'])
                df_cleaned['growth day'] = pd.to_numeric(df_cleaned['growth day'])
                # Convert columns to string
                df_cleaned['farm_name'] = translate(farm)
                df_cleaned = df_cleaned.drop('farm name', axis=1)
                df_cleaned['hatchery'] = df_cleaned['hatchery'].fillna('Unknown').astype(str)

                df_cleaned['line'] = df_cleaned['line'].astype(str)

                try:  # Convert 'date' to datetime
                    df_cleaned['date'] = pd.to_datetime(df_cleaned['date'])
                except:
                    print("error " + farm)

                min_date = df_cleaned['date'].min()
                # Extract the year from the minimum date
                min_year = min_date.year
                # Concatenate the year with the values in the 'folk' column
                df_cleaned['new_flock'] = df_cleaned['flock'].astype(str).apply(lambda x: f"{min_year}{x}")
                df_cleaned['new_flock'] = pd.to_numeric(df_cleaned['new_flock'])
                min_folk = df_cleaned['new_flock'].min()
                farms_new_folk[farm] = min_folk

                # Read existing data from the database
                # Read existing data from the database
                print('now ' + farm)
                existing_data = pd.read_sql(
                    'SELECT [growth_day],[house_number], new_flock, farm_name FROM [dbo].[tmuta]', con=engine)
                existing_data = existing_data.astype(
                    {'growth_day': 'int64', 'house_number': 'int64', 'new_flock': 'int64', 'farm_name': 'str'})

                existing_data.set_index(['growth_day', 'house_number', 'new_flock', 'farm_name'], inplace=True)

                # Ensure df_cleaned has required columns for further processing
                required_columns_df = ['growth day', 'site', 'parent_flock','house_number', 'farm_name', 'hatchery', 'line', 'date',
                                       'mixed_start', 'daily_mortality',
                                       'new_flock']  # Include new_folk in required fields

                missing_columns_df = [col for col in required_columns_df if col not in df_cleaned.columns]

                if missing_columns_df:
                    print(f"Missing columns in df_cleaned for further processing: {missing_columns_df}")
                else:
                    # Creating the DataFrame with specified required columns
                    df = df_cleaned[
                        ['growth day', 'site', 'house_number', 'parent_flock','farm_name', 'hatchery', 'line', 'date', 'mixed_start',
                         'daily_mortality', 'new_flock']]

                    # Convert DataFrame columns to appropriate types
                    df = df.astype(
                        {'growth day': 'int64', 'new_flock': 'int64', 'parent_flock': 'int64','house_number': 'int64', 'farm_name': 'str'})
                    df.set_index(['growth day', 'house_number', 'new_flock', 'farm_name'], inplace=True)
                    # df.set_index(['growth day', 'site', 'house_number', 'farm_name'], inplace=True)

                    # Identify new rows
                    new_rows = df[~df.index.isin(existing_data.index)].reset_index()

                    # Upsert logic: Insert new rows and update existing ones
                    if not new_rows.empty:
                        with engine.begin() as connection:
                            for _, row in new_rows.iterrows():
                                try:
                                    stmt = text("""
                                            MERGE INTO tmuta AS target
                                            USING (SELECT :growth_day AS growth_day, 
                                                          :site AS site, 
                                                          :house_number AS house_number, 
                                                          :parent_flock AS parent_flock, 
                                                          :farm_name AS farm_name, 
                                                          :hatchery AS hatchery, 
                                                          :line AS line, 
                                                          :date AS date, 
                                                          :mixed_start_quantity AS mixed_start_quantity, 
                                                          :daily_mortality AS daily_mortality, 
                                                          :new_flock AS new_flock) AS source
                                            ON target.growth_day = source.growth_day 
                                               AND target.site = source.site 
                                               AND target.house_number = source.house_number 
                                               AND target.parent_flock = source.parent_flock 
                                               AND target.farm_name = source.farm_name 
                                               AND target.hatchery = source.hatchery 
                                               AND target.line = source.line 
                                               AND target.date = source.date
                                            WHEN MATCHED THEN
                                                UPDATE SET target.mixed_start_quantity = source.mixed_start_quantity, 
                                                           target.daily_mortality = source.daily_mortality,
                                                           target.new_flock = source.new_flock  -- Update the new_folk value
                                            WHEN NOT MATCHED BY TARGET THEN
                                                INSERT (growth_day, site, house_number, parent_flock, farm_name, hatchery, line, date, mixed_start_quantity, daily_mortality, new_flock)
                                                VALUES (:growth_day, :site, :house_number, :parent_flock, :farm_name, :hatchery, :line, :date, :mixed_start_quantity, :daily_mortality, :new_flock);
                                        """)

                                    params = {
                                        'growth_day': row['growth day'],
                                        'site': row['site'],
                                        'house_number': row['house_number'],
                                        'parent_flock':  row['parent_flock'],
                                        'farm_name': row['farm_name'],
                                        'hatchery': row['hatchery'],
                                        'line': row['line'],
                                        'date': row['date'],
                                        'mixed_start_quantity': row['mixed_start'],
                                        'daily_mortality': row['daily_mortality'],
                                        'new_flock': row['new_flock']  # Pass new_folk for upsert
                                    }

                                    connection.execute(stmt, params)
                                    print("Success for farm: " + row['farm_name'])
                                except Exception as e:
                                    print(f"An error occurred: {e}")
                        print("Upserted rows successfully.")
                    else:
                        add_flock(translate(farm))
                        print("No new rows to upsert.")
                        # Fetch data and write to MongoDB

def insert_data_to_sql_generic(df, table_name):
    try:
        # Retrieve primary keys
        with engine.connect() as connection:
            stmt = text("""
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
                WHERE TABLE_NAME = :table_name
                AND TABLE_SCHEMA = 'dbo'
                AND OBJECTPROPERTY(OBJECT_ID(CONSTRAINT_SCHEMA + '.' + CONSTRAINT_NAME), 'IsPrimaryKey') = 1;
            """)
            result = connection.execute(stmt, {'table_name': table_name})
            primary_keys = [row[0].lower() for row in result]

        if not primary_keys:
            raise ValueError(f"No primary keys found for table {table_name}.")

        # Normalize column names
        df.columns = df.columns.str.lower()

        # Query existing rows
        query = f"SELECT {', '.join(primary_keys)} FROM {table_name}"
        existing_rows = pd.read_sql_query(query, engine)
        # Normalize column names in both DataFrames
        df.columns = df.columns.str.lower()
        existing_rows.columns = existing_rows.columns.str.lower()

        # Ensure primary keys exist in both DataFrames
        for key in primary_keys:
            if key not in df.columns:
                raise ValueError(f"Primary key '{key}' not found in input DataFrame.")
            if key not in existing_rows.columns:
                raise ValueError(f"Primary key '{key}' not found in existing table.")

        # Ensure data types are consistent
        for key in primary_keys:
            df[key] = df[key].astype(str)  # Convert to string for consistency
            existing_rows[key] = existing_rows[key].astype(str)

        # Debug: Print the DataFrames for inspection
        #print("Input DataFrame (df):")
        #print(df.head())
        #print("Existing Rows:")
        #print(existing_rows.head())

        # Perform deduplication
        deduplicated_df = df.merge(existing_rows, on=primary_keys, how='left', indicator=True)
        deduplicated_df = deduplicated_df[deduplicated_df['_merge'] == 'left_only'].drop(columns=['_merge'])

        # Debug: Print deduplicated DataFrame
        #print("Deduplicated DataFrame:")
        #print(deduplicated_df)

        if deduplicated_df.empty:
            print(f"No new data to insert into {table_name}.")
            return
        else:
            print(f"Data ready for insertion into {table_name}.")

        # Insert rows
        Session = sessionmaker(bind=engine)
        session = Session()
        try:
            for _, row in deduplicated_df.iterrows():
                stmt = text(f"""
                    INSERT INTO {table_name} ([תאריך], teuda, sug_tarovet, [כמות בתעודה], [הפרש], [הפרש משקל], new_flock, farm_name)
                    VALUES (:date, :teuda, :sug_tarovet, :amount, :difference, :weight_diff, :new_flock, :farm_name)
                """)
                try:
                    session.execute(stmt, {
                        'date': row['תאריך'],
                        'teuda': row['teuda'],
                        'sug_tarovet': row['sug_tarovet'],
                        'amount': row['כמות בתעודה'],
                        'difference': row['הפרש'],
                        'weight_diff': row['הפרש משקל'],
                        'new_flock': row['new_flock'],
                        'farm_name': row['farm_name']
                    })
                except IntegrityError:
                    print(f"Duplicate detected: {row.to_dict()}")
                except Exception as row_error:
                    print(f"Error inserting row: {row.to_dict()} | Error: {row_error}")
            session.commit()
            print(f"Data successfully inserted into {table_name}.")
        except Exception as e:
            session.rollback()
            print(f"Error during bulk insert: {e}")
        finally:
            session.close()

    except Exception as e:
        print(f"Error inserting data into {table_name}: {e}")

def insert_data_to_sql_generic1(df, table_name):
    """
    Inserts data from a DataFrame into a SQL table, avoiding duplicates and handling schema dynamically.
    """
    try:
        # Retrieve primary keys
        with engine.connect() as connection:
            stmt = text("""
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
                WHERE TABLE_NAME = :table_name
                AND TABLE_SCHEMA = 'dbo'
                AND OBJECTPROPERTY(OBJECT_ID(CONSTRAINT_SCHEMA + '.' + CONSTRAINT_NAME), 'IsPrimaryKey') = 1;
            """)
            result = connection.execute(stmt, {'table_name': table_name})
            primary_keys = [row[0].lower() for row in result]

        if not primary_keys:
            raise ValueError(f"No primary keys found for table {table_name}.")

        # Normalize column names
        df.columns = df.columns.str.lower()

        # Query existing rows
        query = f"SELECT {', '.join(primary_keys)} FROM {table_name}"
        existing_rows = pd.read_sql_query(query, engine)
        existing_rows.columns = existing_rows.columns.str.lower()

        # Ensure primary keys exist in both DataFrames
        for key in primary_keys:
            if key not in df.columns:
                raise ValueError(f"Primary key '{key}' not found in input DataFrame.")
            if key not in existing_rows.columns:
                raise ValueError(f"Primary key '{key}' not found in existing table.")

        # Deduplication
        deduplicated_df = df.merge(existing_rows, on=primary_keys, how='left', indicator=True)
        deduplicated_df = deduplicated_df[deduplicated_df['_merge'] == 'left_only'].drop(columns=['_merge'])

        if deduplicated_df.empty:
            print(f"No new data to insert into {table_name}.")
            return

        # Insert rows
        Session = sessionmaker(bind=engine)
        session = Session()
        try:
            for _, row in deduplicated_df.iterrows():
                stmt = text(f"""
                    INSERT INTO {table_name} ([farm_name], [sug_tarovet], [teuda], [new_flock], [value], [mivne], [date])
                    VALUES (:farm_name, :sug_tarovet, :teuda, :new_flock, :value, :mivne, :date)
                """)
                try:
                    session.execute(stmt, {
                        'farm_name': row['farm_name'],
                        'sug_tarovet': row['sug_tarovet'],
                        'teuda': row['teuda'],
                        'new_flock': row['new_flock'],
                        'value': row['value'],
                        'mivne': row['mivne'],
                        'date': row['date']
                    })
                except IntegrityError:
                    print(f"Duplicate detected: {row.to_dict()}")
                except Exception as row_error:
                    print(f"Error inserting row: {row.to_dict()} | Error: {row_error}")
            session.commit()
            print(f"Data successfully inserted into {table_name}.")
        except Exception as e:
            session.rollback()
            print(f"Error during bulk insert: {e}")
        finally:
            session.close()

    except Exception as e:
        print(f"Error inserting data into {table_name}: {e}")



def insert_data_to_sql(df, table_name):
    with engine.begin() as connection:
        for _, row in df.iterrows():
            stmt = text("""
                MERGE INTO sivuk AS target
                USING (SELECT :marketing_date AS marketing_date, :house AS house, :receipt AS receipt,
                              :destination AS destination, :marketed_quantity AS marketed_quantity,
                              :averrage_weight AS averrage_weight, :marketed_age AS marketed_age,
                              :farm_name AS farm_name, :new_flock AS new_flock) AS source
                ON target.marketing_date = source.marketing_date AND target.house = source.house
                   AND target.receipt = source.receipt AND target.destination = source.destination
                   AND target.marketed_quantity = source.marketed_quantity AND target.averrage_weight = source.averrage_weight
                   AND target.marketed_age = source.marketed_age AND target.farm_name = source.farm_name
                WHEN MATCHED THEN
                    UPDATE SET marketed_quantity = :marketed_quantity, averrage_weight = :averrage_weight,
                               marketed_age = :marketed_age, farm_name = :farm_name, new_flock = :new_flock
                WHEN NOT MATCHED THEN
                    INSERT (marketing_date, house, receipt, destination, marketed_quantity, averrage_weight, marketed_age, farm_name, new_flock)
                    VALUES (:marketing_date, :house, :receipt, :destination, :marketed_quantity, :averrage_weight, :marketed_age, :farm_name, :new_flock);
            """)

            # Prepare parameters for the SQL command
            params = {
                'marketing_date': row['marketing date'],
                'house': row['house'],
                'receipt': row['receipt'],
                'destination': row['destination'],
                'marketed_quantity': row['marketed quantity'],
                'averrage_weight': row['averrage weight '],
                'marketed_age': row['marketed age'],
                'farm_name': row['farm name'],
                'new_flock': row['new_flock']  # Add new_flock parameter
            }
            connection.execute(stmt, params)


def update_sivuk():
    farms_names = subfolder_names(excel_prod + farms)
    sivuk_results = pd.DataFrame()
    for farm in farms_names:
        try:
            # check if excel file has changed
            print('sivuk ' + farm)
            path = excel_prod + farms + '\\' + farm + excel_middle_name + excel_file_name_finish + farm + excel_end
            data = read_excel(path, sheet_name_sivuk)
            if not data.empty:
                threshold = 5
                # Delete columns with fewer non-null values than the threshold
                data = data.dropna(axis=1, thresh=threshold)
                data = data.dropna(axis=0, thresh=threshold)
                data.columns = data.iloc[0]
                data = data.iloc[1:]
                data = data.reset_index(drop=True)
                data['neto weight '] = pd.to_numeric(data['neto weight '], errors='coerce')
                data = data[data['neto weight '] > 0]
                if not data.empty:
                    new_flock = 'new_flock'  # Define the new flock column name
                    data[new_flock] = farms_new_folk[farm]

                    data['farm name'] = str(translate(farm))
                    sivuk_results = pd.concat([sivuk_results, data], ignore_index=True)

                    # Convert the 'marketing date' column to datetime, allowing for mixed formats
                    sivuk_results['marketing date'] = pd.to_datetime(sivuk_results['marketing date'], errors='coerce')

                    # Format the datetime objects to the desired format (YYYY.MM.DD)
                    sivuk_results['marketing date'] = sivuk_results['marketing date'].dt.strftime('%Y.%m.%d')

                    insert_data_to_sql(sivuk_results, 'sivuk')
        except Exception:
            logging.error(f"Error in farm '{farm}'", exc_info=True)
            print(f"Error in farm '{farm}', details written to logs/sivuk_errors.log")


def update_tarovet():
    #truncate_flock()
    farms_names = subfolder_names(excel_prod + farms)
    tmuta_results = pd.DataFrame()
    for farm in farms_names:
        print('tarovet '+farm)
        # check if excel file has changed
        path = excel_prod + farms + '\\' + farm + excel_middle_name + excel_file_name_finish + farm + excel_end
        data = read_excel(path, sheet_name_tarovet)
        if not data.empty:
            threshold = 5
            # Delete columns with fewer non-null values than the threshold
            data = data.iloc[2:]
            #data = data.dropna(axis=1, thresh=threshold)
            data.columns = data.iloc[0]
            data = data.iloc[1:].reset_index(drop=True)

            data = data.dropna(axis=0, thresh=threshold)

            data['תאריך'] = pd.to_datetime(data['תאריך'], errors='coerce')
            data1 = data[data['תאריך'].notna()]

            date_column = data1[['תאריך']]

            # Identify columns that are numeric by their names
            # Convert column names that are float-like to integers if possible

            new_columns = []
            for col in data1.columns:
                try:
                    # Try converting to float first, then to int if it's a float
                    col_as_int = int(float(col))
                    new_columns.append(col_as_int)
                except (ValueError, TypeError):
                    # If conversion fails, keep the column name as is
                    new_columns.append(col)

            # Update the column names with the new list
            data1.columns = new_columns

            # Now proceed with your numeric_columns check
            numeric_columns = [col for col in data1.columns if col != 'תאריך' and str(col).isnumeric()]

            numeric_columns = [col for col in data1.columns if col != 'תאריך' and str(col).isnumeric()]
            non_numeric_columns = [col for col in data1.columns if col != 'תאריך' and not str(col).isnumeric()]
            data1.rename(columns={data1.columns[1]: "teuda"}, inplace=True)

            data1_numeric = pd.concat([date_column,data1['teuda'],data1['סוג תערובת'], data1[numeric_columns]], axis=1)

            df_long = pd.melt(
                data1_numeric,
                id_vars=["תאריך", "teuda",'סוג תערובת'],  # Include 'teuda' as an id_var
                var_name="mivne",
                value_name="value"
            )

            df_long = df_long.rename(columns={"תאריך": "date"})

            if not data1.empty:
                df_filtered = df_long[df_long["value"].notna()]
                df_filtered['value'] = pd.to_numeric(df_filtered['value'], errors='coerce').astype('Int64')
                new_flock = 'new_flock'  # Define the new flock column name

                try:
                    df_filtered[new_flock] = farms_new_folk[farm]
                except Exception as e:
                    # Optionally, log the error
                    print(f"Error processing farm {farm}: {e}")
                    # Skip the current iteration
                    continue



                df_filtered['farm_name'] = translate(farm)  # Assigning the value of the variable 'farm' to 'farm_name'


                # DataFrame with 'תאריך' and numeric columns

                data1_numeric = pd.concat([date_column, data1[numeric_columns]], axis=1)
                non_numeric_columns[0] = "teuda"
                data1_non_numeric = pd.concat([date_column, data1[non_numeric_columns]], axis=1)
                data1_non_numeric = data1_non_numeric.dropna(axis=1, how='all')
                try:
                    data1_non_numeric['new_flock'] = farms_new_folk[farm]
                except Exception as e:
                    # Optionally, log the error
                    print(f"Error processing farm {farm}: {e}")

                data1_non_numeric['farm_name']= translate(farm)
                data1_non_numeric.columns.values[2] = 'sug_tarovet'
                insert_data_to_sql_generic(data1_non_numeric, 'tarovet')
                df_filtered.columns.values[2] = 'sug_tarovet'
                insert_data_to_sql_generic1( df_filtered, 'tarovet_mivne')
def update_data():
    farms_names = subfolder_names(excel_prod + farms)
    tmuta_results = pd.DataFrame()
    for farm in farms_names:
        # check if excel file has changed
        path = excel_prod + farms + '\\' + farm + excel_middle_name + excel_file_name_finish + farm + excel_end
        data = read_excel(path, sheet_name_tmuta)
        if not data.empty:
            threshold = 5
            # Delete columns with fewer non-null values than the threshold
            data = data.dropna(axis=1, thresh=threshold)
            data = data.dropna(axis=0, thresh=threshold)
            data.columns = data.iloc[0]
            data = data.iloc[1:]
            data = data.reset_index(drop=True)
            data = data[data['mixed daily mortality'] > 0]

            client = MongoClient('mongodb://localhost:27017/')
            db = client['lulim_new']
            collection = db['tmuta']

            # Iterate over DataFrame rows
            for index, row in data.iterrows():
                # Check if the row already exists in the collection
                existing_row = collection.find_one(row.to_dict())
                if existing_row is None:
                    # Insert the row into the collection
                    collection.insert_one(row.to_dict())
                    print("Inserted row:", row.to_dict())
                else:
                    print("Row already exists:", row.to_dict())

            print("Data insertion completed.")

        path = excel_prod + farms + '\\' + farm + excel_middle_name + excel_file_name_finish + farm + excel_end
        data = read_excel(path, sheet_name_sivuk)
        if not data.empty:
            threshold = 5
            # Delete columns with fewer non-null values than the threshold
            data = data.dropna(axis=1, thresh=threshold)
            data = data.dropna(axis=0, thresh=threshold)
            data.columns = data.iloc[0]
            data = data.iloc[1:]
            data = data.reset_index(drop=True)
            data['neto weight '] = pd.to_numeric(data['neto weight '], errors='coerce')
            data = data[data['neto weight '] > 0]
            client = MongoClient('mongodb://localhost:27017/')
            db = client['lulim_new']
            collection = db['sivuk']

            # Iterate over DataFrame rows
            for index, row in data.iterrows():
                # Check if the row already exists in the collection
                existing_row = collection.find_one(row.to_dict())
                if existing_row is None:
                    # Insert the row into the collection
                    collection.insert_one(row.to_dict())
                    print("Inserted row:", row.to_dict())
                else:
                    print("Row already exists:", row.to_dict())

            print("Data insertion completed.")


def update_results():
    # Connect to MongoDB
    client = MongoClient('mongodb://localhost:27017/')
    db = client['lulim_new']  # Replace 'your_database' with your actual database name
    collection = db['tmuta']  # Replace 'your_collection' with your actual collection name

    # Convert MongoDB collection to DataFrame
    cursor = collection.find()

    df = pd.DataFrame(list(cursor))
    # df['date'] = pd.to_datetime(df['date'])

    # Check for missing values
    # print(df.isnull().sum())

    # If there are missing values, handle or remove them as needed

    # Perform the groupby operation after ensuring data type compatibility

    df = df[df['date'] != '#VALUE!']
    # Display the rows with invalid dates

    latest_dates = df.groupby('farm name')['date'].max().reset_index()
    # Group by farm name and find the latest date for each farm
    latest_dates = df.groupby('farm name')['date'].max().reset_index()

    filtered_df = pd.merge(df, latest_dates[['farm name', 'date']], on=['farm name', 'date'])

    filtered_df_date = df.groupby('farm name')['growth day'].max().reset_index()

    filtered_df_date = pd.merge(filtered_df_date, latest_dates[['farm name', 'date']], on=['farm name'])

    filtered_df_date['begin_date'] = pd.to_datetime(filtered_df['date']) - pd.to_timedelta(filtered_df['growth day'],
                                                                                           unit='d')

    filtered_df_date.set_index('farm name', inplace=True)

    # Convert filtered_df to a dictionary
    filtered_dict = filtered_df_date.to_dict(orient='index')

    # df['begin_date'] = ''
    df['begin_date'] = df['farm name'].map(lambda x: filtered_dict.get(x, pd.NaT))
    df['begin_date'] = df['begin_date'].apply(lambda x: x['begin_date'] if isinstance(x, dict) else x)
    df['date'] = pd.to_datetime(df['date'])
    df['begin_date'] = pd.to_datetime(df['begin_date'])

    # Filter the DataFrame where 'date' is greater than or equal to 'begin_date'
    df = df[df['date'] >= df['begin_date']]

    # Iterate over the dictionary and populate the 'begin_date' column in filtered_df
    for farm_name, begin_date in filtered_dict.items():
        df.loc[df['farm name'] == farm_name, 'begin_date'] = begin_date

    result_aggregete = df.groupby('farm name').agg(
        {'mixed daily mortality': 'sum', 'growth day': 'max', 'date': 'max'}). \
        reset_index()
    result_aggregete['hacnasa'] = result_aggregete.apply(
        lambda row: find_hathala(row['farm name'], row['growth day'], df), axis=1)
    result_aggregete['notru_lesivuk'] = result_aggregete['hacnasa'] - result_aggregete['mixed daily mortality']
    result_aggregete['tmuta_iomit'] = result_aggregete.apply(
        lambda row: find_tmuta_iomit(row['farm name'], row['growth day'], df), axis=1)
    write_to_mongo_and_delete(result_aggregete, 'lulim_new', 'tmuta_end')

    client = MongoClient('mongodb://localhost:27017/')
    db = client['lulim_new']  # Replace 'your_database' with your actual database name
    collection = db['sivuk']  # Replace 'your_collection' with your actual collection name

    # Convert MongoDB collection to DataFrame
    cursor = collection.find()

    df = pd.DataFrame(list(cursor))

    # Group by farm name and find the latest date for each farm
    try:
        invalid_date_rows = df[df['marketing date'] == '#VALUE!']
        # Display the rows with invalid dates
        print(invalid_date_rows)
    except KeyError as e:
        print(f"KeyError: {e}")

    latest_dates = df.groupby('farm name')['marketing date'].max().reset_index()

    filtered_df = pd.merge(df, latest_dates[['farm name', 'marketing date']], on=['farm name', 'marketing date'])

    filtered_df_date = df.groupby('farm name')['marketing date'].max().reset_index()

    filtered_df_date = pd.merge(filtered_df_date, latest_dates[['farm name', 'marketing date']], on=['farm name'])

    filtered_df_date['begin_date'] = pd.to_datetime(filtered_df['date']) - pd.to_timedelta(filtered_df['growth day'],
                                                                                           unit='d')

    filtered_df_date.set_index('farm name', inplace=True)

    # Convert filtered_df to a dictionary
    filtered_dict = filtered_df_date.to_dict(orient='index')

    # df['begin_date'] = ''
    df['begin_date'] = df['farm name'].map(lambda x: filtered_dict.get(x, pd.NaT))
    df['begin_date'] = df['begin_date'].apply(lambda x: x['begin_date'] if isinstance(x, dict) else x)
    df['date'] = pd.to_datetime(df['date'])
    df['begin_date'] = pd.to_datetime(df['begin_date'])

    # Filter the DataFrame where 'date' is greater than or equal to 'begin_date'
    df = df[df['date'] >= df['begin_date']]

    # Iterate over the dictionary and populate the 'begin_date' column in filtered_df
    for farm_name, begin_date in filtered_dict.items():
        df.loc[df['farm name'] == farm_name, 'begin_date'] = begin_date

        # Convert the 'marketing date' column to a datetime format
    sivuk_results = df
    sivuk_results['marketing date'] = pd.to_datetime(sivuk_results['marketing date'], format='%d.%m.%y')

    # Format the datetime objects without leading zeros
    sivuk_results['marketing date'] = sivuk_results['marketing date'].dt.strftime('%Y.%m.%d')

    # Create a date object three days ago
    three_days_ago = pd.to_datetime('today') - pd.to_timedelta(days_ago, unit='D')

    # Filter based on the condition
    agg_sivuk_results = sivuk_results[pd.to_datetime(sivuk_results['marketing date']) >= three_days_ago]

    columns_to_drop = ['site', 'flock']
    agg_sivuk_results.drop(columns=columns_to_drop, inplace=True)
    agg_sivuk_results['averrage weight '] = agg_sivuk_results['averrage weight '].fillna(0).round(2)
    agg_sivuk_small = agg_sivuk_results[
        ['farm name', 'marketing date', 'house', 'receipt', 'destination', 'marketed quantity', 'averrage weight ',
         'marketed age']]
    write_to_mongo_and_delete(agg_sivuk_small, 'lulim_new', 'sivuk_end')


def update_views():
    #df_view = pd.read_sql("SELECT * FROM dbo.skila_svuit_highest_grotwh_day;", con=engine)
    df_view2 = pd.read_sql(
        "SELECT [שם חווה],[כמות התחלתית],[תמותה כוללת],[יום גידול],[אחוז תמותה כולל],[יום עדכון אחרון] FROM tmuta14_new ORDER BY [יום עדכון אחרון] desc;",
        con=engine)
    #df_view3 = pd.read_sql("SELECT * FROM dbo.skila_svuit_highest_grotwh_day;", con=engine)
    # Assuming write_to_mongo_and_delete is a defined function
    #write_to_mongo_and_delete(df_view, 'lulim_new', 'tmuta')
    write_to_mongo_and_delete(df_view2, 'lulim_new', 'tmuta14')
    df_view4 = pd.read_sql(
        """
        SELECT 
            [farm_name], 
            [new_flock], 
            [begin_date], 
            [status], 
            [end_date], 
            [begin_quantity], 
            [mortality14], 
            [percent14_tmuta], 
            [total_tarovet], 
            [total_daily_mortality], 
            [marketed_quantity], 
            [total_weight], 
            [marketing_weight], 
            [mesukan_weight], 
            [nezilut_mazon]
        FROM [dbo].[vw_summary_by_flock_no_nulls];
        """,
        con=engine
    )

    # Write the DataFrame to MongoDB and delete it
    write_to_mongo_and_delete(df_view4, 'lulim_new', 'sikum_midgar')

    df_view5 = pd.read_sql(
        "SELECT [farm_name], [new_flock], [house_number],[begin_date], [status], [end_date], "
        "[begin_quantity], [mortality14], [percent14_tmuta] FROM [dbo].[summary_by_flock_mivne];",
        con=engine
    )

    # Write the DataFrame to MongoDB and delete it
    write_to_mongo_and_delete(df_view5, 'lulim_new', 'sikum_midgar_mivne')

    df_view6 = pd.read_sql(
        """
        SELECT [farm_name],
               [min_begin_date],
               [max_end_date],
               [begin_quantity],
               [total_mortality14],
               [total_tarovet],
               [total_marketed_quantity],
               [total_weight],
               [total_marketing_weight],
               [total_daily_mortality],
               [percent14_tmuta],
               [marketing_weight],
               [mesukan_weight],
               [nezilut_mazon],
               [avvarage_sivuk_age],
               [ketzev_gdila_iomi]
        FROM [dbo].[total_summary_2025]
        """,
        con=engine
    )

    # Write the DataFrame to MongoDB and delete it
    write_to_mongo_and_delete(df_view6, 'lulim_new', 'sikum_midgar_summary_2025')


# function that calculates diff between 2 dates




def check():
    from sqlalchemy import text

    with engine.connect() as conn:
        # Begin a transaction
        transaction = conn.begin()
        try:
            # Insert Hebrew data
            insert_query = text("""
                INSERT INTO tarovet ([תאריך], teuda, sug_tarovet, [כמות בתעודה], [הפרש], [הפרש משקל], new_flock, farm_name)
                VALUES ('2024-12-10', 'sh2431581', 1511, 13520, 0, -13520, 20244, N'גוטליב 2');
            """)
            conn.execute(insert_query)

            # Commit the transaction
            transaction.commit()

            # Retrieve data
            select_query = text("SELECT farm_name FROM tarovet WHERE farm_name = N'גוטליב 2';")
            result = conn.execute(select_query)
            for row in result:
                print(row)
        except Exception as e:
            # Roll back in case of an error
            transaction.rollback()
            print(f"An error occurred: {e}")


def job():
    try:
        #check()
        udate_tmuta()
        update_tarovet()
        update_sivuk()
        udate_skila()
        update_views()
        #update_flock()
    except ValueError as e:
        print('bug: ' + e)


def run_program():
    try:
        parser = argparse.ArgumentParser(description="Script to run a job based on environment and command-line arguments")
        #test_sql_connection()
        parser.add_argument("--param1", type=int, help="Param1")
        args = parser.parse_args()

        param1 = args.param1
        print('param 1 input '+str(param1))

        #environment = 'dev'  # You need to set the 'environment' variable somewhere

        if environment == 'dev' or param1 == 0:
            print('now')
            job()

        else:
            if environment == 'prod':
                # Schedule the job to run every 360 minutes (6 hours)
                print('begin prod option')
                job()
                print('finish prod option')
                print('begin prod schedule')
                schedule.every(120).minutes.do(job)
                print('finish prod schedule')

            while True:
                schedule.run_pending()
                time.sleep(1)

    except Exception as e:
        print(f"An error occurred: {e}")
        print("Restarting the program...")
        run_program()

if __name__ == "__main__":
    run_program()